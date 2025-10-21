"""
Boeing 737NG/737MAX 油尺计算器
基于油尺读数和滚转角计算燃油量
"""

import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import time
import json
import os

# 页面配置
st.set_page_config(
    page_title="Boeing 737NG/MAX 油尺计算器",
    page_icon="✈️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 自定义CSS样式（优化版本）
st.markdown("""
<style>
    /* 主要样式优化 */
    .main-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 2rem;
        border-radius: 15px;
        color: white;
        text-align: center;
        margin-bottom: 2rem;
        box-shadow: 0 8px 32px rgba(102, 126, 234, 0.3);
        animation: fadeInDown 0.8s ease-out;
    }
    
    .main-header h1 {
        font-size: 2.5rem;
        font-weight: 700;
        margin-bottom: 0.5rem;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.3);
    }
    
    .main-header p {
        font-size: 1.2rem;
        opacity: 0.9;
        margin: 0;
    }
    
    /* 性能指标样式 */
    .performance-metrics {
        background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
        padding: 1rem;
        border-radius: 10px;
        color: white;
        margin-bottom: 1rem;
        text-align: center;
    }
    
    /* 快速加载指示器 */
    .loading-indicator {
        display: flex;
        align-items: center;
        justify-content: center;
        padding: 1rem;
        background: #f8f9fa;
        border-radius: 10px;
        margin: 1rem 0;
    }
    
    .spinner {
        border: 3px solid #f3f3f3;
        border-top: 3px solid #667eea;
        border-radius: 50%;
        width: 20px;
        height: 20px;
        animation: spin 1s linear infinite;
        margin-right: 10px;
    }
    
    @keyframes spin {
        0% { transform: rotate(0deg); }
        100% { transform: rotate(360deg); }
    }
    
    /* 其他现有样式保持不变 */
    .info-box {
        background: linear-gradient(135deg, #74b9ff 0%, #0984e3 100%);
        padding: 1.5rem;
        border-radius: 12px;
        color: white;
        margin: 1rem 0;
        box-shadow: 0 4px 15px rgba(116, 185, 255, 0.3);
        animation: slideInLeft 0.6s ease-out;
    }
    
    .warning-box {
        background: linear-gradient(135deg, #fdcb6e 0%, #e17055 100%);
        padding: 1.5rem;
        border-radius: 12px;
        color: white;
        margin: 1rem 0;
        box-shadow: 0 4px 15px rgba(253, 203, 110, 0.3);
        animation: slideInRight 0.6s ease-out;
    }
    
    .error-box {
        background: linear-gradient(135deg, #fd79a8 0%, #e84393 100%);
        padding: 1.5rem;
        border-radius: 12px;
        color: white;
        margin: 1rem 0;
        box-shadow: 0 4px 15px rgba(253, 121, 168, 0.3);
        animation: shake 0.5s ease-in-out;
    }
    
    .success-box {
        background: linear-gradient(135deg, #00b894 0%, #00a085 100%);
        padding: 1.5rem;
        border-radius: 12px;
        color: white;
        margin: 1rem 0;
        box-shadow: 0 4px 15px rgba(0, 184, 148, 0.3);
        animation: bounceIn 0.6s ease-out;
    }
    
    .metric-container {
        background: linear-gradient(135deg, #a29bfe 0%, #6c5ce7 100%);
        padding: 1.5rem;
        border-radius: 12px;
        color: white;
        text-align: center;
        margin: 0.5rem 0;
        box-shadow: 0 4px 15px rgba(162, 155, 254, 0.3);
        transition: transform 0.3s ease, box-shadow 0.3s ease;
    }
    
    .metric-container:hover {
        transform: translateY(-5px);
        box-shadow: 0 8px 25px rgba(162, 155, 254, 0.4);
    }
    
    .feature-card {
        background: white;
        padding: 1.5rem;
        border-radius: 12px;
        box-shadow: 0 4px 15px rgba(0,0,0,0.1);
        margin: 1rem 0;
        border-left: 4px solid #667eea;
        transition: transform 0.3s ease, box-shadow 0.3s ease;
    }
    
    .feature-card:hover {
        transform: translateY(-3px);
        box-shadow: 0 8px 25px rgba(0,0,0,0.15);
    }
    
    .quick-button {
        background: linear-gradient(135deg, #00cec9 0%, #55a3ff 100%);
        color: white;
        border: none;
        padding: 0.8rem 1.5rem;
        border-radius: 25px;
        font-weight: 600;
        cursor: pointer;
        transition: all 0.3s ease;
        margin: 0.5rem;
        box-shadow: 0 4px 15px rgba(0, 206, 201, 0.3);
    }
    
    .quick-button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 20px rgba(0, 206, 201, 0.4);
    }
    
    .data-table {
        background: white;
        border-radius: 12px;
        overflow: hidden;
        box-shadow: 0 4px 15px rgba(0,0,0,0.1);
        margin: 1rem 0;
    }
    
    .progress-container {
        background: #f8f9fa;
        padding: 1rem;
        border-radius: 10px;
        margin: 1rem 0;
        text-align: center;
    }
    
    .status-text {
        font-weight: 600;
        color: #2d3436;
        margin-bottom: 0.5rem;
    }
    
    /* 动画效果 */
    @keyframes fadeInDown {
        from { opacity: 0; transform: translateY(-30px); }
        to { opacity: 1; transform: translateY(0); }
    }
    
    @keyframes slideInLeft {
        from { opacity: 0; transform: translateX(-30px); }
        to { opacity: 1; transform: translateX(0); }
    }
    
    @keyframes slideInRight {
        from { opacity: 0; transform: translateX(30px); }
        to { opacity: 1; transform: translateX(0); }
    }
    
    @keyframes bounceIn {
        0% { opacity: 0; transform: scale(0.3); }
        50% { opacity: 1; transform: scale(1.05); }
        70% { transform: scale(0.9); }
        100% { opacity: 1; transform: scale(1); }
    }
    
    @keyframes shake {
        0%, 100% { transform: translateX(0); }
        10%, 30%, 50%, 70%, 90% { transform: translateX(-5px); }
        20%, 40%, 60%, 80% { transform: translateX(5px); }
    }
    
    /* 响应式设计 */
    @media (max-width: 768px) {
        .main-header h1 { font-size: 2rem; }
        .main-header p { font-size: 1rem; }
        .feature-card { margin: 0.5rem 0; padding: 1rem; }
    }
</style>
""", unsafe_allow_html=True)

# ========= Excel 数据源工具函数 =========

def load_excel_by_model(model_choice: str):
    """根据机型加载对应的 Excel 数据文件并返回工作簿对象"""
    excel_path = os.path.join(os.getcwd(), "737MAX.xlsx" if model_choice == "737MAX" else "737NG.xlsx")
    if not os.path.exists(excel_path):
        return None, None, f"未找到 Excel 数据文件：{excel_path}"
    try:
        wb = load_workbook(excel_path, data_only=True)
        return excel_path, wb, None
    except Exception as e:
        return excel_path, None, f"加载 Excel 工作簿失败：{e}"

def get_sheetnames_for_sticks(wb):
    """获取与油尺相关的工作表名称（优先匹配包含“油尺”的表）"""
    names = wb.sheetnames
    stick_sheets = [n for n in names if "油尺" in n]
    return stick_sheets if stick_sheets else names

def get_roll_angles(ws, wing: str):
    """根据机翼选择获取横滚角列表：Right Wing 用第1行（14->6），Left Wing 用第2行（6->14）"""
    header_row = 1 if wing == "Right Wing" else 2
    roll_angles = []
    # 从第3列开始读取到最后一列
    for c in range(3, ws.max_column + 1):
        val = ws.cell(row=header_row, column=c).value
        if val is None:
            continue
        try:
            roll_angles.append(float(val))
        except Exception:
            # 可能存在字符串如 "6°"，尝试去除非数字字符
            try:
                roll_angles.append(float(str(val).replace("°", "").strip()))
            except Exception:
                continue
    return roll_angles

def get_pitch_values(ws):
    """从第3行开始读取第二列的俯仰角，并返回去重后的有序列表（按出现顺序）"""
    seen = []
    for r in range(3, ws.max_row + 1):
        val = ws.cell(row=r, column=2).value
        if val is None:
            continue
        try:
            fv = float(val)
        except Exception:
            try:
                fv = float(str(val).replace("°", "").strip())
            except Exception:
                continue
        if fv not in seen:
            seen.append(fv)
    return seen

def get_scales_for_pitch(ws, pitch: float):
    """在给定俯仰角的连续数据块中，收集第一列的油尺刻度"""
    scales = []
    # 遍历并匹配俯仰角
    for r in range(3, ws.max_row + 1):
        pv = ws.cell(row=r, column=2).value
        if pv is None:
            continue
        try:
            pvf = float(pv)
        except Exception:
            try:
                pvf = float(str(pv).replace("°", "").strip())
            except Exception:
                continue
        if abs(pvf - pitch) < 1e-9:
            sv = ws.cell(row=r, column=1).value
            if sv is None:
                continue
            try:
                scales.append(float(sv))
            except Exception:
                try:
                    scales.append(float(str(sv).strip()))
                except Exception:
                    continue
    return scales

def find_row_for_pitch_scale(ws, pitch: float, scale: float):
    """查找满足俯仰角与油尺刻度的行号（精确匹配），未找到返回 None"""
    for r in range(3, ws.max_row + 1):
        pv = ws.cell(row=r, column=2).value
        sv = ws.cell(row=r, column=1).value
        if pv is None or sv is None:
            continue
        try:
            pvf = float(pv)
        except Exception:
            try:
                pvf = float(str(pv).replace("°", "").strip())
            except Exception:
                continue
        try:
            svf = float(sv)
        except Exception:
            try:
                svf = float(str(sv).strip())
            except Exception:
                continue
        if abs(pvf - pitch) < 1e-9 and abs(svf - scale) < 1e-9:
            return r
    return None

def find_col_for_roll(ws, wing: str, roll: float):
    """根据机翼选择与横滚角找到对应列号"""
    header_row = 1 if wing == "Right Wing" else 2
    for c in range(3, ws.max_column + 1):
        val = ws.cell(row=header_row, column=c).value
        if val is None:
            continue
        try:
            rf = float(val)
        except Exception:
            try:
                rf = float(str(val).replace("°", "").strip())
            except Exception:
                continue
        if abs(rf - roll) < 1e-9:
            return c
    return None

def lookup_fuel(ws, row_idx: int, col_idx: int):
    """读取指定单元格的燃油值（磅），返回 float 或 None"""
    if row_idx is None or col_idx is None:
        return None
    val = ws.cell(row=row_idx, column=col_idx).value
    if val is None:
        return None
    try:
        return float(val)
    except Exception:
        try:
            return float(str(val).strip())
        except Exception:
            return None

def interpolate_fuel_by_scale(ws, pitch: float, wing: str, roll: float, scale_input: float):
    """
    针对给定俯仰角与机翼的横滚角列，按油尺刻度进行线性插值：
    - 若刻度恰好匹配表中的某行，直接返回该单元格燃油值
    - 若刻度落在相邻刻度之间，返回线性插值值
    - 若超出范围或数据不足，返回 (None, reason)

    返回: (fuel_value, info) 其中 info 指示是否为插值/精确匹配/错误信息
    """
    col_idx = find_col_for_roll(ws, wing, float(roll))
    if col_idx is None:
        return None, "未找到对应的横滚角列"

    # 收集指定俯仰角的 (scale, fuel) 对
    pairs = []
    for r in range(3, ws.max_row + 1):
        pv = ws.cell(row=r, column=2).value
        if pv is None:
            continue
        try:
            pvf = float(pv)
        except Exception:
            try:
                pvf = float(str(pv).replace("°", "").strip())
            except Exception:
                continue
        if abs(pvf - float(pitch)) > 1e-9:
            continue

        sv = ws.cell(row=r, column=1).value
        fv = ws.cell(row=r, column=col_idx).value
        if sv is None or fv is None:
            continue
        # 解析刻度与燃油值
        try:
            svf = float(sv)
        except Exception:
            try:
                svf = float(str(sv).strip())
            except Exception:
                continue
        try:
            fvf = float(fv)
        except Exception:
            try:
                fvf = float(str(fv).strip())
            except Exception:
                continue
        pairs.append((svf, fvf))

    if not pairs:
        return None, "未收集到可用于插值的刻度-燃油数据"

    # 按刻度排序
    pairs.sort(key=lambda x: x[0])

    # 精确匹配检查
    for svf, fvf in pairs:
        if abs(svf - float(scale_input)) < 1e-9:
            return fvf, "精确匹配"

    # 范围检查与线性插值
    smin, smax = pairs[0][0], pairs[-1][0]
    s = float(scale_input)
    if s < smin or s > smax:
        return None, f"刻度 {s} 超出有效范围 [{smin}, {smax}]"

    # 找到相邻区间
    for i in range(len(pairs) - 1):
        x0, y0 = pairs[i]
        x1, y1 = pairs[i + 1]
        if x0 <= s <= x1:
            if abs(x1 - x0) < 1e-12:
                return y0, "相邻刻度相同，返回下界值"
            t = (s - x0) / (x1 - x0)
            y = y0 + (y1 - y0) * t
            return y, f"线性插值：[{x0},{y0}] 与 [{x1},{y1}]，t={t:.3f}"

    # 理论上不应到达此处
    return None, "插值失败：未定位到相邻区间"

# ========= 燃油密度换算与校正 =========
NOMINAL_DENSITY_LB_PER_GAL = 6.76  # 名义密度 (lbs/US gallon)
KG_PER_LB = 0.45359237
L_PER_GAL = 3.785411784

def density_to_lbs_per_gal(value: float, unit: str) -> float:
    """将密度转换为 lbs/US gallon。支持 'lbs/US gallon' 与 'kg/L'"""
    if value is None:
        return None
    unit = (unit or '').strip().lower()
    if unit in ['lbs/us gallon', 'lbs/gal', 'lb/gal', 'lb/us gallon']:
        return float(value)
    elif unit in ['kg/l', 'kg per l', 'kg\u002fl']:
        # lbs/gal = (kg/L * L/gal) / kg/lb
        return float(value) * L_PER_GAL / KG_PER_LB
    else:
        # 未知单位，回退为原值（假定用户输入的是 lbs/gal）
        return float(value)

def compute_correction_factor(actual_density_lbs_per_gal: float) -> float:
    """校正系数 = 名义密度 / 实际密度"""
    if actual_density_lbs_per_gal is None or actual_density_lbs_per_gal <= 0:
        return None
    return NOMINAL_DENSITY_LB_PER_GAL / float(actual_density_lbs_per_gal)

def main():
    # 初始化变量
    fuel_amount = None
    
    # 设置页面样式
    st.markdown("""
    <style>
    .main-header {
        background: linear-gradient(120deg, #f8fafc 0%, #e2e8f0 100%);
        border-radius: 15px;
        padding: 25px 35px;
        margin: 10px 0 35px 0;
        box-shadow: 0 4px 15px rgba(0, 0, 0, 0.05),
                   0 10px 30px rgba(0, 0, 0, 0.02);
        border: 1px solid rgba(255, 255, 255, 0.5);
    }
    .main-header h1 {
        color: #1e40af;
        font-size: 2.4em;
        margin-bottom: 20px;
        font-weight: 700;
        letter-spacing: -0.5px;
        text-shadow: 1px 1px 2px rgba(0, 0, 0, 0.05);
        display: flex;
        align-items: center;
        gap: 12px;
    }
    .main-header h1 span {
        font-size: 1.8em;
        line-height: 1;
        margin-right: 5px;
    }
    .main-header p {
        color: #334155;
        font-size: 1.15em;
        line-height: 1.7;
        margin: 12px 0;
        padding-left: 2px;
    }
    .main-header p:last-child {
        color: #64748b;
        font-size: 0.95em;
        font-style: italic;
        border-top: 1px solid rgba(148, 163, 184, 0.2);
        padding-top: 15px;
        margin-top: 15px;
        padding-left: 2px;
    }
    .developer-info {
        margin-top: 20px;
        background: linear-gradient(to right, #f1f5f9, #e2e8f0);
        padding: 12px;
        border-radius: 8px;
        font-size: 0.85em;
        color: #64748b;
        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.05);
        border: 1px solid rgba(255, 255, 255, 0.7);
    }
    .developer-info p {
        margin: 0;
        line-height: 1.6;
    }
    .developer-info .dev-name {
        color: #475569;
        font-weight: 500;
    }
    .developer-info .dev-date {
        color: #94a3b8;
        font-size: 0.95em;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # 主标题
    st.markdown("""
    <div class="main-header">
        <h1><span>✈️</span>Boeing 737NG/MAX 油尺计算器</h1>
        <p>本系统基于Boeing 737NG/MAX FUEL MEASURING STICK MANUAL提供精确的燃油量计算功能。仅供学习和研究使用，请勿用于商业用途。</p>
        <p>免责声明：本工具仅供参考，实际操作请严格遵循官方手册和相关规定。</p>
    </div>
    """, unsafe_allow_html=True)
    
    # 侧边栏添加机型选择，并按机型加载对应 Excel 数据文件
    st.sidebar.header("🛩️ 机型选择")
    model_choice = st.sidebar.selectbox(
        "选择机型",
        options=["737NG", "737MAX"],
        index=0,
        help="选择飞机机型以加载对应的数据表"
    )
    # 加载 Excel 工作簿
    excel_path, wb, error = load_excel_by_model(model_choice)
    if error:
        st.markdown(f"""
        <div class=\"error-box\">
            <h3>❌ 加载失败</h3>
            <p>{error}</p>
            <p>请确保 737NG.xlsx / 737MAX.xlsx 存在于应用目录。</p>
        </div>
        """, unsafe_allow_html=True)
        return
    stick_sheets = get_sheetnames_for_sticks(wb)
    
    
    # 侧边栏 - 输入参数
    with st.sidebar:
        st.header("📝 输入参数")
        st.caption(f"当前机型：{model_choice} | 数据文件：{os.path.basename(excel_path)}")

        # 油尺（机翼）选择：仅 Left / Right
        st.subheader("🛢️ 选择油尺")
        selected_wing = st.selectbox(
            "选择油尺",
            options=["Left Wing", "Right Wing"],
            index=0,
            help="选择机翼：Right Wing 使用第一行滚转角（14→6），Left Wing 使用第二行（6→14）"
        )

        # 油尺编号（工作表）选择
        stick_sheet = st.selectbox(
            "选择油尺编号",
            options=stick_sheets,
            index=0,
            help="选择对应的油尺工作表"
        )
        ws = wb[stick_sheet]

        # 飞行参数
        st.subheader("🎛️ 参数")

        # 俯仰角候选（来自该工作表）
        pitch_values = get_pitch_values(ws)
        default_pitch_idx = 0
        if pitch_values:
            default_pitch_idx = int(np.argmin(np.abs(np.array(pitch_values) - 0.0)))
        pitch = st.selectbox(
            "俯仰角 (度)",
            options=pitch_values if pitch_values else [0.0],
            index=default_pitch_idx if pitch_values else 0,
            help="选择俯仰角度"
        )

        # 横滚角候选（根据机翼选择的表头行）
        roll_angles = get_roll_angles(ws, selected_wing)
        roll = st.selectbox(
            "横滚角 (度)",
            options=roll_angles if roll_angles else [6.0, 7.0, 8.0, 9.0, 10.0, 11.0, 12.0, 13.0, 14.0],
            index=len(roll_angles)//2 if roll_angles else 4,
            help="选择横滚角度"
        )

        # 针对选定俯仰角提供刻度范围，并改为输入框+线性插值
        scale_options = get_scales_for_pitch(ws, float(pitch))
        if scale_options:
            dynamic_scale_min = float(min(scale_options))
            dynamic_scale_max = float(max(scale_options))
        else:
            dynamic_scale_min = 0.0
            dynamic_scale_max = 100.0
        # 油尺读数默认空值，使用文本输入并在后续解析
        scale_input = st.text_input(
            "油尺读数",
            value="",
            placeholder=f"请输入刻度值（{dynamic_scale_min:.1f} ~ {dynamic_scale_max:.1f}），支持线性插值",
            help=f"可输入任意刻度值；若不在表格刻度上，将在相邻值间线性插值。有效范围：{dynamic_scale_min:.1f} ~ {dynamic_scale_max:.1f}"
        )
        # 将字符串解析为浮点数；空串保持为 None
        scale = None
        if scale_input.strip() != "":
            try:
                scale = float(scale_input.strip())
            except ValueError:
                st.markdown(
                    """
                    <div class=\"error-box\">
                        <h4>❌ 输入错误</h4>
                        <p>油尺读数必须为数字，例如 5.4 或 12.0。</p>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
        # 如果已解析为数值，做范围预检查
        if scale is not None and not (dynamic_scale_min <= scale <= dynamic_scale_max):
            st.markdown(
                f"""
                <div class=\"error-box\">
                    <h4>❌ 输入错误</h4>
                    <p>油尺读数超出有效范围：{dynamic_scale_min:.1f} ~ {dynamic_scale_max:.1f}</p>
                </div>
                """,
                unsafe_allow_html=True,
            )

        # 在侧边栏底部添加开发人员信息
        st.markdown("")  # 添加一些空间
        st.markdown("")
        st.markdown("""
        <div class="developer-info">
            <p class="dev-name">👨‍💻 开发人员：王康业</p>
            <p class="dev-date">📅 更新日期：2025年10月21日</p>
        </div>
        """, unsafe_allow_html=True)

    
    col1, col2 = st.columns([2, 1])

    with col2:
        st.subheader("⚖️ 密度校正")
        enable_density_correction = st.checkbox(
            "启用燃油密度校正", value=True,
            key="density_correction_main",
            help="当实际燃油密度 ≠ 6.76 lbs/US gallon 时，按照手册进行校正。"
        )
        density_unit = st.selectbox(
            "密度单位",
            options=["lbs/US gallon", "kg/L"],
            index=0,
            help="选择实际燃油密度的单位"
        )
        # 根据单位给出一个合理的默认值（名义密度）
        default_density_value = NOMINAL_DENSITY_LB_PER_GAL if density_unit == "lbs/US gallon" else (NOMINAL_DENSITY_LB_PER_GAL * KG_PER_LB / L_PER_GAL)
        actual_density_input = st.number_input(
            "实际燃油密度",
            min_value=0.1,
            max_value=20.0 if density_unit == "lbs/US gallon" else 2.0,
            value=float(f"{default_density_value:.3f}"),
            step=0.001,
            format="%.3f",
            help="示例：更致密 6.85 lbs/US gallon (≈0.810 kg/L)；更稀 6.55 lbs/US gallon (≈0.774 kg/L)"
        )
        # 开始计算按钮移至右侧模块
        start_calculation = st.button("▶️ 开始计算")

    with col1:
        st.subheader("🧮 计算结果")
        if start_calculation:
            # 基于 Excel 的查表计算（支持刻度线性插值）
            validation_errors = []
            # 校验油尺读数是否已填写
            if scale is None:
                validation_errors.append("请输入有效的油尺读数。")
                fuel_amount_raw, interp_info = None, None
            else:
                # 插值计算
                fuel_amount_raw, interp_info = interpolate_fuel_by_scale(
                    ws, float(pitch), selected_wing, float(roll), float(scale)
                )
            if fuel_amount_raw is None:
                validation_errors.append(interp_info or "插值失败")
            if validation_errors:
                for error in validation_errors:
                    st.markdown(f"""
                    <div class=\"error-box\">
                        <h4>❌ 输入错误</h4>
                        <p>{error}</p>
                    </div>
                    """, unsafe_allow_html=True)
            else:
                start_calc_time = time.time()
                corrected_fuel_load = fuel_amount_raw  # 基于名义密度的表值（CFL），可能为插值结果
                calc_time = time.time() - start_calc_time

                if corrected_fuel_load is not None:
                    # 若启用密度校正，则计算实际燃油量 AFL = CFL / CF
                    if enable_density_correction:
                        actual_density_lbs_per_gal = density_to_lbs_per_gal(actual_density_input, density_unit)
                        correction_factor = compute_correction_factor(actual_density_lbs_per_gal)
                        if correction_factor is None:
                            st.markdown("""
                            <div class=\"error-box\">
                                <h3>❌ 密度校正错误</h3>
                                <p>请输入有效的实际燃油密度（>0）。</p>
                            </div>
                            """, unsafe_allow_html=True)
                        else:
                            actual_fuel_load = corrected_fuel_load / correction_factor
                            st.markdown(f"""
                            <div class=\"success-box\">
                                <h3>✅ 计算成功（含密度校正）</h3>
                                <p><strong>表查得燃油量（默认密度 6.76）:</strong> {corrected_fuel_load:.2f} 磅</p>
                                <p><strong>查表说明:</strong> {interp_info}</p>
                                <p><strong>校正系数:</strong> CF = 6.76 / {actual_density_lbs_per_gal:.3f} = {correction_factor:.3f}</p>
                                <h2 style=\"font-size: 2.0rem; margin: 0.5rem 0;\">实际燃油量（校正后）: {actual_fuel_load:.2f} 磅</h2>
                                <p><strong>计算时间:</strong> {calc_time*1000:.2f} 毫秒</p>
                                <p><strong>油尺工作表:</strong> {stick_sheet} | <strong>机翼:</strong> {selected_wing}</p>
                                <p><strong>俯仰角:</strong> {pitch}° | <strong>横滚角:</strong> {roll}° | <strong>读数:</strong> {scale}</p>
                            </div>
                            """, unsafe_allow_html=True)
                    else:
                        st.markdown(f"""
                        <div class=\"success-box\">
                            <h3>✅ 计算成功</h3>
                            <h2 style=\"font-size: 2.0rem; margin: 0.5rem 0;\">{corrected_fuel_load:.2f} 磅</h2>
                            <p><strong>查表说明:</strong> {interp_info}</p>
                            <p><strong>计算时间:</strong> {calc_time*1000:.2f} 毫秒</p>
                            <p><strong>油尺工作表:</strong> {stick_sheet} | <strong>机翼:</strong> {selected_wing}</p>
                            <p><strong>俯仰角:</strong> {pitch}° | <strong>横滚角:</strong> {roll}° | <strong>读数:</strong> {scale}</p>
                        </div>
                        """, unsafe_allow_html=True)
                else:
                    st.markdown("""
                    <div class=\"error-box\">
                        <h3>❌ 计算失败</h3>
                        <p>未能从工作表中读取到有效燃油值，请检查输入参数或工作表数据。</p>
                    </div>
                    """, unsafe_allow_html=True)
        else:
            st.info("点击右侧的‘开始计算’按钮以执行计算。")
    

if __name__ == "__main__":
    main()
